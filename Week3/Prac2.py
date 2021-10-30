#Homework 3 ex2

#Author: Yanbo Liang
#Date: 2021.10.26

import csv

header = ['Power1', 'Power2', 'Power3', 'Power4', 'Power5']
f = open("ex2csv.csv","w", newline='')
#create writer
writer = csv.writer(f)
writer.writerow(header)

data = ['','','','','']
for k in range(1,101):
    for i in range(1,6):
        data[i-1] = '%d^%d' % (k,i)
        if i ==5:
            writer.writerow(data)

for k in range(1,101):
    for i in range(1,6):
        data[i-1] = '%d^%d' % (101-k,6-i)
        if i ==5:
            writer.writerow(data)
#### CSV creation completed - without using panda

## ---- xlsx creation ----
from openpyxl import load_workbook
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
i = 1
for x in header:
    sheet.cell(row=1, column = i).value = x
    i +=1

for k in range(1,101):
    for i in range(1,6):
        sheet.cell(row=k+1, column=i).value = '%d^%d' % (k,i)

for k in range(1,101):
    for i in range(1,6):
        sheet.cell(row=k+101, column=i).value = '%d^%d' % (101-k,6-i)


wb.save(filename="ex2excel.xlsx")

#### XLSX creation completed -- without panda

import pandas

df1 = pandas.DataFrame(columns = ['Power1', 'Power2', 'Power3', 'Power4', 'Power5'])



for i in range(1,6):
    p1 = list()
    for k in range(1,101):
        p1.append('%d^%d' % (k,i))

    for k in range(1,101):
        p1.append('%d^%d' % (101-k,6-i))

    df1['Power'+str(i)] = p1
df1.to_csv('ex2pandacsv.csv')

#### Panda CSV done

#### XLSX panda creation

import pandas


df1 = pandas.DataFrame(columns = ['Power1', 'Power2', 'Power3', 'Power4', 'Power5'])



for i in range(1,6):
    p1 = list()
    for k in range(1,101):
        p1.append('%d^%d' % (k,i))

    for k in range(1,101):
        p1.append('%d^%d' % (101-k,6-i))

    df1['Power'+str(i)] = p1
df1.to_excel('ex2pandaexcel.xlsx')
